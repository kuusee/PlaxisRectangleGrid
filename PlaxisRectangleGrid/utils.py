#!/usr/bin/env python
from typing import Dict, NoReturn, Union

from datetime import datetime
import re

import pandas as pd
import numpy as np

from openpyxl.styles import (PatternFill, Font, Border,
                             Alignment, Side, NamedStyle)

"""
Пакет содержит стили для оформления таблиц, функции для предобработки данных из
excel файла, создания таблиц и записи значений в них.
"""
FILL = PatternFill(start_color='e8fdfb', end_color='e8fdfb', fill_type='solid')
FONT = Font(b=True, size=12)
ALIGNMENT = Alignment(horizontal='center', vertical='center')
BORDER = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))

STYLE_HEADER = NamedStyle(name='header')
STYLE_HEADER.font = FONT
STYLE_HEADER.fill = FILL
STYLE_HEADER.border = BORDER
STYLE_HEADER.alignment = ALIGNMENT

STYLE_CELL = NamedStyle(name='cell')
STYLE_CELL.border = BORDER
STYLE_CELL.alignment = ALIGNMENT

STYLE_ERROR = NamedStyle(name='error')
STYLE_ERROR.border = BORDER
STYLE_ERROR.alignment = ALIGNMENT
STYLE_ERROR.fill = PatternFill(start_color='CD5C5C',
                               end_color='CD5C5C',
                               fill_type='solid')

STYLE_SUCCESS = NamedStyle(name='success')
STYLE_SUCCESS.border = BORDER
STYLE_SUCCESS.alignment = ALIGNMENT
STYLE_SUCCESS.fill = PatternFill(start_color='90EE90',
                                 end_color='90EE90',
                                 fill_type='solid')


class SheetInfoStyles(object):
    """
    Класс содержит стили оформления листа журнала операций для класса SheetInfo
    """
    main_cell = NamedStyle(name='main_cell')
    main_cell.border = BORDER
    main_cell.alignment = ALIGNMENT
    main_cell.font = Font(b=True, size=14)

    grid_pt_font = Font(b=True, size=11)

    errors = STYLE_ERROR
    success = STYLE_SUCCESS
    cell = STYLE_CELL
    header = STYLE_HEADER


def markup_excel(sheet,
                 coord_x: np.ndarray,
                 coord_z: np.ndarray,
                 style=STYLE_HEADER) -> NoReturn:
    """
    Создание и оформление таблицы в excel файле.

    :param sheet: лист в excel файле
    :param coord_x: названия столбцов шапки таблицы ([-10, -5, 0, 5, ....])
    :param coord_z: названия строк таблицы ([0, -5, -10, -15, ....])
    :param style: стиль оформления ячеек openpyxl.styles.NamedStyle
    :return: NoReturn
    """

    # Оформление ячейки 1, 1
    sheet.cell(row=1, column=1, value='z|x')
    sheet.cell(row=1, column=1).style = style
    sheet.row_dimensions[1].height = 15
    sheet.column_dimensions['A'].width = 10

    # Оформление шапки столбцов
    for idx, x in enumerate(coord_x, 2):
        column = sheet.cell(row=1, column=idx)
        col_letter = column.column_letter  # получаем букву столбца
        column.value = x
        column.style = style
        sheet.column_dimensions[col_letter].width = 10

    # Оформление шапки строк
    for idx, z in enumerate(coord_z, 2):
        row = sheet.cell(row=idx, column=1)
        row.value = z
        row.style = style
        sheet.row_dimensions[idx].height = 15


def write_excel(sheet,
                row: int,
                column: int,
                value: Union[float, int, str],
                style=STYLE_CELL) -> NoReturn:
    """
    Функция записи значения в ячейку excel таблицы. Запись значения
    типов int, float производится в десятичном формате '0.000000';
    типа str в текстовом формате.

    :param sheet: лист в excel файле
    :param row: номер ячейки для записи
    :param column: номер столбца для записи
    :param value: записываемое значение
    :param style: стиль оформления ячеек openpyxl.styles.NamedStyle
    :return: NoReturn
    """

    cell = sheet.cell(row=row, column=column)
    cell.value = value
    cell.style = style
    if isinstance(value, str):
        cell.number_format = '@'  # текстовый формат
    else:
        cell.number_format = '0.000000'  # десятичный формат


def preprocessing_data(filepath: str,
                       cols: Dict[str, tuple]
                       ) -> Dict[str, np.ndarray]:
    """
    Функция для предобработки данных из excel файла. Извлекает название листов
    и данные из заданных столбцов. Удаляет дубликаты.

    :param filepath: название файла, (например, "100x60x30.xls")
    :param cols: словарь с номера столбцов, вида {название_листа: номера_стлб}
    (например, {'x': (3, 5, 7), 'y': (4, 5, 6)})
    :return: словарь вида {название_листа: массив_данных}
    """

    load_file = pd.read_excel(filepath, sheet_name=None)  # загрузка файла
    # создаем словарь из непустых листов. Ключ-название листа, значение-массив
    # данных
    data = {key.lower(): value for key, value in load_file.items()
            if len(load_file[key])}

    for key, value in data.items():
        # создаем массив данных из заданных столбцов
        points = value.iloc[:, [*cols[key]]]
        points.columns = ['X', 'Z', 'uY']
        points = points.to_numpy()

        data[key] = np.unique(points, axis=0)  # сортируем и удаляем дубликаты
    return data


def parse_filename(filename: str) -> tuple:
    return tuple(map(lambda x: int(x), re.findall("\d+", filename)))


def output_filepath(filename: str, postfix: str) -> str:
    """
    Создает название для выходного файла.

    :param filename: название файла, (например, "100x60x30.xls")
    :param postfix: постфикс, (например, "_CALC_")
    :return: "100x60x30_CALC_.xls"
    """
    path_split = filename.rsplit('.', maxsplit=1)
    return path_split[0] + postfix + '.' + path_split[1]


def filepath_txt(path: str) -> str:
    """
    Генерирует имя файла в виде "%d_%m_%Y__%H_%M_%S.txt"
    например, "21_01_2023__19_37_45.txt". Генерирует путь до него.
    САМ ФАЙЛ НЕ СОЗДАЕТСЯ.
    :param path: путь к папке где создаем файл, например, "../data"
    :return: f'{path}/{datetime.today()}.txt'
    """
    dt_time = datetime.today().strftime("%d_%m_%Y__%H_%M_%S")
    return f'{path}/{dt_time}.txt'


def log_to_string(filepath: str, sheet_name: str, fields: dict) -> str:
    """
    Создает строку для записи в файл txt.

    :param filepath: абсолютный путь до файла, (например, "data/100x60x30.xls")
    :param sheet_name: название листа для которого создаем лог, (например, "x")
    :param fields: словарь с записываемыми данными {'Points': 100,'Success': 0}
    :return: строка "filepath время_записи\nsheet_name\n"".join(fields)\n\n"
    """
    string_1 = filepath + '\t' + datetime.today().strftime("%d.%m.%Y %H:%M:%S")
    string_2 = f'Sheet: {sheet_name}'
    string_3 = "".join([f'{field}: {val}\t' for field, val in fields.items()])
    string_4 = '\n\n'
    return "\n".join([string_1, string_2, string_3, string_4])
