#!/usr/bin/env python

import itertools

import tkinter as tk
from tkinter import filedialog

import numpy as np
from openpyxl import Workbook

from PlaxisRectangleGrid import classes, utils, calc_func as calc


PATH_INPUT = './data'  # путь до папки с файлами
OUTPUT_FILENAME_POSTFIX = '_CALC_'  # постфикс для файла вывода

# названия листов и номера колонок с данными (X, Z, u_Y) входного файла
COLUMNS_COORD = {
    'x': (3, 5, 7),
    'y': (4, 5, 6),
    'default': (3, 5, 7)
}

# шаг сетки в зависимости от глубины Z
STEP_Z = {
    -10: -0.5,
    -15: -1.,
    -22: -1.5,
    -30: -2,
}

EPS_X = 1  # размер окрестности поиска по X
EPS_Z = 1  # размер окрестности поиска по Z

INFO_SHEETS = True  # создаем листы в excel для записи операций


if __name__ == '__main__':
    root = tk.Tk()
    root.withdraw()

    filenames = filedialog.askopenfilenames(initialdir=PATH_INPUT)

    # генерируем имя текстового файла для логов
    file_txt = utils.filepath_txt(PATH_INPUT)

    for filepath in filenames:
        # Создает название для выходного файла
        out_file = utils.output_filepath(filepath, OUTPUT_FILENAME_POSTFIX)

        # создаем файл xls
        xls = Workbook()

        # предобработка данных
        data = utils.preprocessing_data(filepath, COLUMNS_COORD)
        print(filepath)

        for key, value in data.items():
            """
            Проходим в цикле по предобработанному массиву
            key - название листа
            value - массив точек вида [[X, Z, u_Y], ... ...]
            """

            # Определяем границы для построения прямоугольной сетки
            length = np.abs(value[:, 0].max()) + np.abs(value[:, 0].min())
            depth = value[:, 1].min()  # тут верхняя всегда равна 0

            # генерируем узлы прямоугольной сетки
            grid = classes.RectangleGrid(length, depth, STEP_Z)

            # создаем лист для записи перемещений и оформляем его
            sheet = xls.create_sheet(key)
            utils.markup_excel(sheet, grid.grid_x, grid.grid_z)

            # в цикле проходим по всем узлам сетки и вычисляем перемещения
            # в них по ближайщим существующим точкам
            for idxs, point in grid.journal_pts.items():
                row, col = idxs[0], idxs[1]  # координаты для ячейки листа

                # ищем ближайшие точки в заданной окрестности
                nearby_pts, nearby_dist = point.nearby_points(
                    value,
                    eps_x=grid.step_x,
                    eps_z=-grid.step_z
                )

                mean_u_y = []

                try:
                    if len(nearby_pts):

                        # проверяем на взаимное положение точки относительно
                        # точки
                        command, pts = point.point4point2d(nearby_pts,
                                                           nearby_dist)

                        if command == 'point':
                            u_y = pts[0, 2]
                            mean_u_y.append(u_y)

                        elif len(nearby_pts) == 2:

                            # проверяем взаимное расположение точки и прямой,
                            # заданной двумя точками
                            command, pts = point.point4lines2d(nearby_pts)

                            if command == 'line':
                                # вычисляем перемещение для узла point
                                u_y = calc.interpolate(point.coords,
                                                       nearby_pts)
                                mean_u_y.append(u_y)

                        elif len(nearby_pts) > 2:
                            """
                            Если ближайщих точек 3-4 то перемещение для узла 
                            можно вычислить методом пересечения прямой 
                            проходящей через узел point перпендикулярной 
                            плоскости (XoZ) и плоскостиобразованной 3 точками. 
                            Если точек больше 3х - создаем наборы по 3 точки из
                            всех точек, проверяем как расположен узел point 
                            относительно внутренней области треугольника 
                            образованного набором из 3х точек.
                            
                            Если узел лежит на ребре треугольника, то 
                            перемещения вычисляем как для точки лежащей на 
                            проекции линии на плоскость XoZ(интерполяция);
                            
                            Если узел лежит внутри - вычисляем перемещения 
                            методом описанным выше и добавляем в список 
                            перемещений.
                            
                            Если узел лежит снаружи - переходим к следующему 
                            набору.
                            
                            Далее вычисляем среднее перемещение - 
                            оно и будет конечным значением.
                            """
                            sets_point = itertools.combinations(nearby_pts, 3)

                            for set_ in sets_point:

                                command, pts = point.point4triangle2d(set_)

                                if command == 'line':
                                    u_y = calc.interpolate(point.coords, pts)
                                    mean_u_y.append(u_y)
                                    break

                                elif command == 'inside':
                                    u_y = calc.intersect(point.coords, pts)
                                    mean_u_y.append(u_y)
                except Exception:
                    print('Warning')

                if len(mean_u_y):
                    point.u_y = np.mean(mean_u_y)

                    # записываем результат в ячейку excel файла
                    utils.write_excel(sheet, row, col, point.u_y)

                    # обновляем значение успешных узлов
                    grid.main_info['Success'] += 1

                else:
                    point.u_y = None
                    utils.write_excel(sheet, row, col, 'NULL',
                                      style=utils.STYLE_ERROR)

                    # обновляем значение ошибок узлов
                    grid.main_info['Errors'] += 1

            if INFO_SHEETS:
                # создаем лист для записи операций
                sheet_info = xls.create_sheet('info_' + key)
                sht_info = classes.SheetInfo(sheet_info, grid.main_info)
                for point in grid.journal_pts.values():
                    sht_info.write_journal(point)  # записываем лог операций

            # записываем лог в файл
            with open(file_txt, "a") as logs:
                log_string = utils.log_to_string(filepath, key, grid.main_info)
                logs.write(log_string)

        xls.remove(xls['Sheet'])
        xls.save(out_file)
